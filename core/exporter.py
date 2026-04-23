from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from core.balancer import result_to_rows
from data.models import AssignmentResult, MappingRecord, ValidationIssue


class ReportExporter:
    def export_excel(
        self,
        path: str | Path,
        result: AssignmentResult,
        mappings: list[MappingRecord],
        issues: list[ValidationIssue],
    ) -> None:
        workbook = Workbook()
        result_sheet = workbook.active
        result_sheet.title = "排产结果"

        headers = [
            "工位",
            "员工",
            "工位负荷",
            "process_no",
            "component",
            "description",
            "standard_process_id",
            "standard_process_name",
            "standard_time",
            "effective_time",
        ]
        result_sheet.append(headers)
        for row in result_to_rows(result):
            result_sheet.append([row.get(header, "") for header in headers])

        mapping_sheet = workbook.create_sheet("映射确认")
        mapping_sheet.append(["工序描述", "LLM候选", "最终标准工序", "最终工序ID", "置信度", "人工确认", "原因"])
        for record in mappings:
            mapping_sheet.append(
                [
                    record.process_description,
                    record.llm_skill_name,
                    record.final_skill_name,
                    record.final_process_id,
                    record.confidence,
                    "是" if record.human_approved else "否",
                    record.reason,
                ]
            )

        issue_sheet = workbook.create_sheet("导入异常")
        issue_sheet.append(["级别", "对象", "行", "字段", "说明"])
        for issue in issues:
            issue_sheet.append([issue.severity, issue.entity, issue.row_index, issue.field, issue.message])

        metrics_sheet = workbook.create_sheet("指标")
        metrics_sheet.append(["平衡率", result.metrics.balance_rate])
        metrics_sheet.append(["节拍", result.metrics.cycle_time])
        metrics_sheet.append(["工位数", result.metrics.num_stations])
        metrics_sheet.append(["总有效工时", result.metrics.total_effective_time])
        metrics_sheet.append(["总搬运距离", result.metrics.total_distance])
        for warning in result.warnings:
            metrics_sheet.append(["提示", warning])

        workbook.save(path)

    def export_pdf(self, path: str | Path, result: AssignmentResult, mappings: list[MappingRecord]) -> None:
        pdf_path = Path(path)
        font_name = _register_chinese_font()
        pdf = canvas.Canvas(str(pdf_path), pagesize=A4)
        width, height = A4
        y = height - 48

        pdf.setFont(font_name, 14)
        pdf.drawString(48, y, "裤子车缝生产线车位排产报告")
        y -= 28
        pdf.setFont(font_name, 10)
        pdf.drawString(48, y, f"平衡率：{result.metrics.balance_rate * 100:.1f}%")
        y -= 18
        pdf.drawString(48, y, f"节拍：{result.metrics.cycle_time:.2f} 分钟")
        y -= 18
        pdf.drawString(48, y, f"总搬运距离：{result.metrics.total_distance:.1f}")
        y -= 26
        pdf.drawString(48, y, "工位分配：")
        y -= 18

        for station in result.stations:
            if y < 80:
                pdf.showPage()
                pdf.setFont(font_name, 10)
                y = height - 48
            pdf.drawString(54, y, f"{station.station_id} {station.employee_name} 负荷 {station.load_time:.2f} 分钟")
            y -= 16
            for process in station.assigned_processes[:4]:
                pdf.drawString(72, y, f"{process['process_no']} {process['standard_process_name']} / {process['description'][:26]}")
                y -= 14

        pdf.showPage()
        pdf.setFont(font_name, 10)
        pdf.drawString(48, height - 48, "映射确认摘要")
        y = height - 72
        for record in mappings[:45]:
            if y < 60:
                pdf.showPage()
                pdf.setFont(font_name, 10)
                y = height - 48
            pdf.drawString(
                48,
                y,
                f"{record.process_description[:28]} -> {record.final_skill_name or '待确认'} ({record.confidence:.2f})",
            )
            y -= 14
        pdf.save()


def _register_chinese_font() -> str:
    candidates = [
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simsun.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
    ]
    for font in candidates:
        if font.exists():
            try:
                pdfmetrics.registerFont(TTFont("Chinese", str(font)))
                return "Chinese"
            except Exception:
                continue
    return "Helvetica"
